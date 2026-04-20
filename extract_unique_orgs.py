"""
Unique Organisation Extractor
===============================
Reads the affiliations Excel file (output of fetch_affiliations.py) and
produces a deduplicated list of unique organisations, normalised using the
ROR (Research Organization Registry) API.

The problem:
  Raw Crossref/OpenAlex affiliations include departments, faculties, cities,
  states — making the same institution appear in many different forms:
    - "Centre for Mental Health, Faculty of Health, Arts & Design,
       Swinburne University of Technology, Melbourne, VIC, Australia"
    - "Swinburne University of Technology Melbourne Victoria Australia"
    - "Swinburne University of Technology"
  All of these should resolve to "Swinburne University of Technology".

Strategy:
  1. Collect all raw affiliation strings from the input.
  2. For each unique raw string, query the ROR affiliation API which is
     specifically designed to parse messy affiliation strings and return
     the best-matching institution with a confidence score.
  3. Group raw strings by their resolved ROR institution.
  4. Output: one row per unique institution with ROR ID, country, type,
     all raw variants, and the authors/DOIs linked to it.

Resilience: JSON cache, periodic saves, retry, graceful Ctrl+C.

Setup:
  Same config.ini as the other scripts in this pipeline.

Usage:
  python extract_unique_orgs.py <affiliations.xlsx> [output.xlsx] [config.ini]
"""

import sys
import re
import json
import time
import html
import configparser
from pathlib import Path
from collections import defaultdict
import requests
import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill

ROR_API = "https://api.ror.org/v2/organizations"
SCRIPT_DIR = Path(__file__).resolve().parent


# ── Config ──────────────────────────────────────────────────────────────

def load_config(config_path=None):
    if config_path is None:
        config_path = SCRIPT_DIR / "config.ini"
    else:
        config_path = Path(config_path).resolve()

    if not config_path.exists():
        print(f"ERROR: Config file not found at {config_path}")
        sys.exit(1)

    config = configparser.ConfigParser()
    config.read(config_path)
    email = config.get("crossref", "email", fallback=None)
    if not email or email.strip() == "your_email@example.com":
        print(f"ERROR: Please set your real email in {config_path}")
        sys.exit(1)

    return {
        "email": email.strip(),
        "delay": config.getfloat("crossref", "delay", fallback=1),
        "save_every": config.getint("crossref", "save_every", fallback=50),
        "max_retries": config.getint("crossref", "max_retries", fallback=3),
    }


# ── Cache ───────────────────────────────────────────────────────────────

def load_cache(cache_path):
    if cache_path.exists():
        with open(cache_path, "r", encoding="utf-8") as f:
            cache = json.load(f)
        print(f"  Loaded cache: {len(cache)} entries from {cache_path.name}")
        return cache
    return {}


def save_cache(cache, cache_path):
    with open(cache_path, "w", encoding="utf-8") as f:
        json.dump(cache, f, ensure_ascii=False, indent=2)


# ── Affiliation string parsing ──────────────────────────────────────────

def decode_and_split_affiliations(raw_cell_value):
    """
    Decode HTML entities FIRST (so &amp; becomes & and doesn't break
    semicolon splitting), then split by semicolons.
    """
    if not raw_cell_value:
        return []

    text = str(raw_cell_value)
    # Decode HTML entities: &amp; -> &, &lt; -> <, etc.
    text = html.unescape(text)
    # Split by semicolons
    parts = [p.strip() for p in text.split(';')]
    # Filter empties and placeholder values
    skip = {"[No affiliation found]", ""}
    return [p for p in parts if p and p not in skip]


# ── ROR API ─────────────────────────────────────────────────────────────

def extract_org_from_ror_item(item):
    """Extract structured org info from a ROR API result item."""
    org = item.get("organization", {})

    # Extract primary display name
    names = org.get("names", [])
    display_name = ""
    for n in names:
        if "ror_display" in n.get("types", []):
            display_name = n.get("value", "")
            break
    if not display_name:
        for n in names:
            if "label" in n.get("types", []):
                display_name = n.get("value", "")
                break
    if not display_name and names:
        display_name = names[0].get("value", "")

    # Extract location
    locations = org.get("locations", [])
    country = ""
    city = ""
    if locations:
        geonames = locations[0].get("geonames_details", {})
        country = geonames.get("country_name", "")
        city = geonames.get("name", "")

    # Extract type
    org_types = org.get("types", [])
    org_type = org_types[0] if org_types else ""

    return {
        "ror_id": org.get("id", ""),
        "name": display_name,
        "country": country,
        "city": city,
        "type": org_type,
        "score": item.get("score", 0),
        "chosen": item.get("chosen", False),
    }


def query_ror_affiliation(raw_string, session, max_retries=3):
    """
    Use ROR's affiliation endpoint to resolve a raw affiliation string
    to a canonical institution.
    """
    url = ROR_API
    params = {"affiliation": raw_string}

    for attempt in range(1, max_retries + 1):
        try:
            resp = session.get(url, params=params, timeout=30)
            if resp.status_code == 429:
                wait = min(2 ** attempt * 5, 60)
                print(f"    Rate limited. Waiting {wait}s (retry {attempt}/{max_retries})...")
                time.sleep(wait)
                continue
            resp.raise_for_status()
            data = resp.json()
            items = data.get("items", [])
            if not items:
                return None

            # Prefer the "chosen" result (ROR's best match)
            for item in items:
                if item.get("chosen", False):
                    return extract_org_from_ror_item(item)

            # No "chosen" — use top result if score >= 0.8
            top = items[0]
            if top.get("score", 0) >= 0.8:
                return extract_org_from_ror_item(top)

            return None

        except requests.exceptions.RequestException as e:
            if attempt < max_retries:
                wait = 2 ** attempt
                print(f"    Error: {e}. Retrying in {wait}s ({attempt}/{max_retries})...")
                time.sleep(wait)
            else:
                print(f"    Failed after {max_retries} retries: {e}")
                return None
    return None


# ── Main ────────────────────────────────────────────────────────────────

def main(input_file, output_file=None, config_path=None):
    cfg = load_config(config_path)
    print(f"Config: email={cfg['email']}, delay={cfg['delay']}s")

    # Resolve paths
    input_path = Path(input_file)
    if not input_path.is_absolute():
        if not input_path.exists():
            fallback = SCRIPT_DIR / input_path
            if fallback.exists():
                input_path = fallback
    input_path = input_path.resolve()

    if not input_path.exists():
        print(f"ERROR: Input file not found: {input_path}")
        sys.exit(1)

    if output_file is None:
        output_path = input_path.parent / f"{input_path.stem}_unique_orgs.xlsx"
    else:
        output_path = Path(output_file).resolve()

    cache_path = output_path.parent / f"{input_path.stem}_ror_cache.json"

    print(f"Input:  {input_path}")
    print(f"Output: {output_path}")

    ror_cache = load_cache(cache_path)

    # --- Read input ---
    wb = openpyxl.load_workbook(input_path)
    ws = wb.active

    headers = {cell.value: cell.column for cell in ws[1]}
    name_col = headers.get("Author_Name")
    affs_col = headers.get("Affiliations")
    dois_col = headers.get("DOIs")

    if affs_col is None:
        print("ERROR: Could not find 'Affiliations' column.")
        sys.exit(1)

    # --- Collect raw affiliations with their authors and DOIs ---
    raw_aff_data = defaultdict(lambda: {"authors": set(), "dois": set()})

    for row in range(2, ws.max_row + 1):
        affs_val = ws.cell(row=row, column=affs_col).value
        author = ws.cell(row=row, column=name_col).value if name_col else ""
        dois_val = ws.cell(row=row, column=dois_col).value if dois_col else ""

        if not affs_val:
            continue

        author = str(author).strip() if author else ""
        # DOIs are also semicolon-separated but won't have HTML entities
        dois = [d.strip() for d in str(dois_val).split(';') if d.strip()] if dois_val else []

        # Decode HTML entities BEFORE splitting by semicolons
        affiliations = decode_and_split_affiliations(affs_val)

        for aff in affiliations:
            raw_aff_data[aff]["authors"].add(author)
            raw_aff_data[aff]["dois"].update(dois)

    print(f"\nFound {len(raw_aff_data)} unique raw affiliation strings.")

    # --- Query ROR for each raw affiliation ---
    session = requests.Session()
    session.headers.update({
        "User-Agent": f"OrgExtractor/1.0 (mailto:{cfg['email']})",
        "Accept": "application/json",
    })

    to_fetch = [aff for aff in raw_aff_data if aff not in ror_cache]
    print(f"{len(raw_aff_data) - len(to_fetch)} cached, {len(to_fetch)} to fetch from ROR.\n")

    fetched = 0
    try:
        for raw_aff in to_fetch:
            fetched += 1
            display = raw_aff[:70] + "..." if len(raw_aff) > 70 else raw_aff
            # Encode safely for Windows consoles that can't handle all Unicode
            display = display.encode("ascii", errors="replace").decode("ascii")
            print(f"[{fetched}/{len(to_fetch)}] {display}")

            result = query_ror_affiliation(raw_aff, session, cfg["max_retries"])
            if result:
                ror_cache[raw_aff] = result
            else:
                ror_cache[raw_aff] = {
                    "name": raw_aff, "ror_id": "", "country": "",
                    "city": "", "type": "", "score": 0, "chosen": False
                }

            if fetched % cfg["save_every"] == 0:
                save_cache(ror_cache, cache_path)
                print(f"  >> Cache saved ({fetched}/{len(to_fetch)})")

            if fetched < len(to_fetch):
                time.sleep(cfg["delay"])

    except KeyboardInterrupt:
        print(f"\n>> Interrupted! Saving cache ({fetched} fetched)...")
        save_cache(ror_cache, cache_path)
        print("Re-run to continue.")
        sys.exit(0)

    save_cache(ror_cache, cache_path)

    # --- Group by resolved organisation ---
    org_groups = defaultdict(lambda: {
        "name": "",
        "ror_id": "",
        "country": "",
        "city": "",
        "type": "",
        "raw_variants": set(),
        "authors": set(),
        "dois": set(),
        "best_score": 0,
    })

    for raw_aff, data in raw_aff_data.items():
        resolved = ror_cache.get(raw_aff, {})
        ror_id = resolved.get("ror_id", "")
        resolved_name = resolved.get("name", raw_aff)

        # Use ROR ID as grouping key if available, otherwise normalised name
        if ror_id:
            key = ror_id
        else:
            key = re.sub(r'\s+', ' ', resolved_name.lower().strip())

        group = org_groups[key]

        score = resolved.get("score", 0)
        if score > group["best_score"] or not group["name"]:
            group["name"] = resolved_name
            group["ror_id"] = ror_id
            group["country"] = resolved.get("country", "")
            group["city"] = resolved.get("city", "")
            group["type"] = resolved.get("type", "")
            group["best_score"] = score

        group["raw_variants"].add(raw_aff)
        group["authors"].update(data["authors"])
        group["dois"].update(data["dois"])

    print(f"\nResolved to {len(org_groups)} unique organisations.")

    # --- Sort alphabetically ---
    sorted_orgs = sorted(org_groups.values(), key=lambda g: g["name"].lower())

    # --- Stats ---
    with_ror = sum(1 for g in sorted_orgs if g["ror_id"])
    without_ror = len(sorted_orgs) - with_ror
    print(f"  With ROR ID:    {with_ror}")
    print(f"  Without ROR ID: {without_ror}")

    # --- Write output ---
    out_wb = openpyxl.Workbook()
    out_ws = out_wb.active
    out_ws.title = "Unique Organisations"

    header_font = Font(name="Arial", bold=True, color="FFFFFF")
    header_fill = PatternFill("solid", fgColor="4472C4")
    header_align = Alignment(horizontal="left", vertical="center")

    col_headers = [
        "Organisation_Name", "ROR_ID", "Country", "City", "Type",
        "Raw_Variants", "Author_Count", "Authors", "DOI_Count", "DOIs"
    ]
    for col_idx, hdr in enumerate(col_headers, 1):
        cell = out_ws.cell(row=1, column=col_idx, value=hdr)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_align

    data_font = Font(name="Arial")
    wrap_align = Alignment(wrap_text=True, vertical="top")

    for row_idx, org in enumerate(sorted_orgs, 2):
        out_ws.cell(row=row_idx, column=1, value=org["name"]).font = data_font
        out_ws.cell(row=row_idx, column=2, value=org["ror_id"]).font = data_font
        out_ws.cell(row=row_idx, column=3, value=org["country"]).font = data_font
        out_ws.cell(row=row_idx, column=4, value=org["city"]).font = data_font
        out_ws.cell(row=row_idx, column=5, value=org["type"]).font = data_font

        variants = sorted(org["raw_variants"] - {org["name"]})
        cell = out_ws.cell(row=row_idx, column=6, value="; ".join(variants) if variants else "")
        cell.font = data_font
        cell.alignment = wrap_align

        authors = sorted(org["authors"] - {""})
        out_ws.cell(row=row_idx, column=7, value=len(authors)).font = data_font
        cell = out_ws.cell(row=row_idx, column=8, value="; ".join(authors))
        cell.font = data_font
        cell.alignment = wrap_align

        dois = sorted(org["dois"])
        out_ws.cell(row=row_idx, column=9, value=len(dois)).font = data_font
        cell = out_ws.cell(row=row_idx, column=10, value="; ".join(dois))
        cell.font = data_font
        cell.alignment = wrap_align

    out_ws.column_dimensions['A'].width = 40
    out_ws.column_dimensions['B'].width = 35
    out_ws.column_dimensions['C'].width = 18
    out_ws.column_dimensions['D'].width = 18
    out_ws.column_dimensions['E'].width = 15
    out_ws.column_dimensions['F'].width = 60
    out_ws.column_dimensions['G'].width = 14
    out_ws.column_dimensions['H'].width = 50
    out_ws.column_dimensions['I'].width = 12
    out_ws.column_dimensions['J'].width = 60
    out_ws.freeze_panes = "A2"
    out_ws.auto_filter.ref = f"A1:J{out_ws.max_row}"

    out_wb.save(output_path)
    print(f"\nDone. Output saved to: {output_path}")


if __name__ == "__main__":
    if len(sys.argv) < 2:
        print("Usage: python extract_unique_orgs.py <affiliations.xlsx> [output.xlsx] [config.ini]")
        sys.exit(1)
    in_arg = sys.argv[1]
    out_arg = sys.argv[2] if len(sys.argv) > 2 else None
    cfg_arg = sys.argv[3] if len(sys.argv) > 3 else None
    main(in_arg, out_arg, cfg_arg)